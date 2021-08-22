import pyautogui
import pyperclip
import openpyxl
import time
start = time.time()
wb = openpyxl.load_workbook(r'e:\桌面\当日数据汇总.xlsx', data_only=True,)
sheet = wb['数据']
max_row = sheet.max_row
print('最大行数：',max_row)
pyautogui.PAUSE = 0.7
pyautogui.FAILSAFE = True #要停止程序的话，将鼠标指针移动到左上角
cash_r = 0
for column in list(sheet.columns)[0]:
    cash_r +=1 
    if column.value == '1001.01':
        break
    elif cash_r == max_row:
        cash_r = max_row + 1 
print('现金起始行数：',cash_r)          
pyautogui.alert(text='请将金蝶窗口左上角对齐屏幕的左上角，并打开凭证管理界面，准备好后点确认', title='准备', button='确认')
pyautogui.click(327,126)#点击金蝶窗口
def bank_part():#录入银行收款
    print('\t开始录入银行收款凭证：')
    pyautogui.mouseDown(36,76)
    pyautogui.mouseUp(36,76)
    time.sleep(5)
    # pyautogui.press(['\n']*7) # 光标移动到第一行摘要区域
    pyautogui.click(145,280)
    for rows in range(2,cash_r):# 从第2行开始逐行向下运行到最大行
        if sheet.cell(rows,2).value is None:#判断是否是收款业务
            continue
        pyperclip.copy(sheet.cell(rows,4).value) # 复制摘要
        pyautogui.hotkey('ctrl', 'v') # 粘贴摘要
        pyautogui.press('\n') 
        pyautogui.typewrite(sheet.cell(rows,1).value) #输入银行科目代码
        pyautogui.press('\n') 
        pyautogui.typewrite(str(sheet.cell(rows,5).value)) #输入借方金额
        pyautogui.press(['\n','\n','left']) 
        pyperclip.copy(sheet.cell(rows,4).value)# 复制摘要
        pyautogui.hotkey('ctrl','v')# 粘贴摘要
        pyautogui.press('\n') #回车
        pyautogui.typewrite('1131') #输入应收账款代码
        pyautogui.press('\n') #回车
        pyautogui.typewrite(sheet.cell(rows,2).value)#输入客户代码
        pyautogui.press(['\n','\n','\n','=','\n','\n','left']) 
    pyautogui.press('f12') # 保存凭证
    time.sleep(8)
    print('\t录入完毕')
def cash_part():#录入现金收款
    print('\t开始录入现金凭证:')
    pyautogui.mouseDown(36,76)
    pyautogui.mouseUp(36,76)
    time.sleep(5)
    # pyautogui.press(['\n']*7) # 光标移动到第一行摘要区域
    pyautogui.click(145,280)
    for rows in range(cash_r,max_row+1):# 从现金记录的第1行开始逐行向下运行到最大行
        if sheet.cell(rows,2).value is None:#判断是否是收款业务
            continue
        pyperclip.copy(sheet.cell(rows,4).value) # 复制摘要
        pyautogui.hotkey('ctrl', 'v') # 粘贴摘要
        pyautogui.press('\n') 
        pyautogui.typewrite('1131') #输入应收账款代码
        pyautogui.press('\n') 
        pyautogui.typewrite(sheet.cell(rows,2).value)#输入客户代码
        pyautogui.press(['\n','\n','\n']) 
        pyautogui.typewrite(str(sheet.cell(rows,5).value)) #输入贷方金额
        pyautogui.press(['\n','\n','left']) 
    pyautogui.press(['up','up','\n'])# 光标移动到第一行
    pyautogui.click(357, 76) # 鼠标点击插入
    pyautogui.click(357, 76) # 鼠标点击插入
    pyperclip.copy('收现') # 复制摘要
    pyautogui.hotkey('ctrl', 'v') # 粘贴摘要
    pyautogui.press('\n') 
    pyautogui.typewrite('1001.01') #输入现金代码
    pyautogui.press('\n') 
    pyautogui.press('=')
    pyautogui.press('f12') # 保存凭证
    time.sleep(4)  
    print('\t录入完毕')
def report_part():#资金日报录入
    print(' \t开始录入资金日报表，请等待...')
    sheet2 = wb['日金额合计']
    s_name = time.strftime('%y.%m.%d', time.localtime())
    wb_1 = openpyxl.load_workbook(r'\\Pc-20180314ijns\移动硬盘\2日报表\{}月\{}\加纳资金日报表2020.{}.xlsx'.format(s_name[3:5],s_name[3:],s_name[3:5]))
    day_sheet = wb_1[s_name]
    list1=[2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21]
    list2=[6,7,8,11,12,13,14,15,16,17,18,19,20,21,22,25,26,27,28,29]
    list3=[]
    for i in list1:
        list3.append([sheet2.cell(i,3).value,sheet2.cell(i,4).value])
    a = 0
    for h in list2:
        day_sheet.cell(h,5).value = list3[a][0]
        day_sheet.cell(h,6).value = list3[a][1]
        a +=1
    list44 = [2,3,5,6,7,8,9,10,12,16,17,18,19,20]
    list55 = [6,7,11,12,13,14,15,16,18,22,25,26,27,28]
    for g in range(14):
        day_sheet['m{}'.format(str(list55[g]))] = sheet2.cell(list44[g],5).value
    day_sheet['F34'] = sheet2.cell(24,4).value  #客户交款
    day_sheet['E34'] = sheet2.cell(27,4).value  #开票
    day_sheet['F40'] = sheet2.cell(27,2).value  #wp销售
    day_sheet['G40'] = sheet2.cell(27,3).value  #WP其他
    day_sheet['F40'] = sheet2.cell(28,2).value  #SOL销售
    day_sheet['G40'] = sheet2.cell(28,3).value  #SOL其他
    wb_1.save(r'\\Pc-20180314ijns\移动硬盘\2日报表\{}月\{}\加纳资金日报表2020.{}.xlsx'.format(s_name[3:5],s_name[3:],s_name[3:5]))
    wb.close()
    wb_1.close()
    print('\t录入完毕')
def main():
    choice = pyautogui.confirm(text='请选择录入内容', title='提示', buttons=['默认', '银行', '现金','资金报表'])
    if choice =='默认':
        if cash_r > max_row:
            bank_part()
            report_part()
            print('仅银行&资金报表 录入完毕')
        elif cash_r == 2: 
            cash_part()
            report_part()     
            print('仅现金&资金报表 录入完毕')
        else:
            bank_part()
            cash_part()
            report_part()
            print('银行&现金&资金报表 录入完毕')
    elif choice == '银行':
        bank_part()
        print('银行录入完毕')
    elif choice == '现金':
        cash_part()
        print('现金录入完毕')
    elif choice == '资金报表':
        report_part() 
main()
end = time.time()
result='已完成，耗时:{}秒'.format(round(end-start))
pyautogui.alert(text=result, title='提示', button='OK')

